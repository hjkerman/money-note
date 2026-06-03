from __future__ import annotations

from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, PatternFill, Side


CURRENT_SHEET = "당월 기록"
ARCHIVE_SHEET = "전체 기록(본인)"


@dataclass
class ParsedWorkbook:
    current_entries: list[dict[str, Any]]
    archive_rows: list[dict[str, Any]]
    panels: list[dict[str, Any]]
    settings: dict[str, str]
    labels: dict[str, str]


def parse_amount(value: Any, cached_value: Any = None) -> tuple[float | None, str | None]:
    if isinstance(value, str) and value.startswith("="):
        return _to_float(cached_value), None
    return _to_float(value), None


def _to_float(value: Any) -> float | None:
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def normalize_date_label(value: Any) -> tuple[str | None, str | None]:
    if value is None:
        return None, None
    if isinstance(value, datetime):
        day = value.date()
        return day.isoformat(), f"{day:%Y.%m.%d}."
    if isinstance(value, date):
        return value.isoformat(), f"{value:%Y.%m.%d}."

    text = str(value).strip()
    for fmt in ("%Y.%m.%d.", "%Y.%m.%d", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            day = datetime.strptime(text, fmt).date()
            return day.isoformat(), text
        except ValueError:
            pass
    return None, text


def import_workbook(path: Path) -> ParsedWorkbook:
    formulas = load_workbook(path, data_only=False)
    cached = load_workbook(path, data_only=True)
    current = formulas[CURRENT_SHEET]
    current_cached = cached[CURRENT_SHEET]
    archive = formulas[ARCHIVE_SHEET]
    archive_cached = cached[ARCHIVE_SHEET]

    current_entries = _parse_current_entries(current, current_cached)
    panels = _parse_current_panels(current, current_cached)
    settings = _parse_current_settings(current, current_cached)
    labels = _parse_labels(current, archive)
    archive_rows = _parse_archive_rows(archive, archive_cached)
    return ParsedWorkbook(
        current_entries=current_entries,
        archive_rows=archive_rows,
        panels=panels,
        settings=settings,
        labels=labels,
    )


def _parse_current_entries(ws: Any, cached_ws: Any) -> list[dict[str, Any]]:
    entries: list[dict[str, Any]] = []
    last_date: tuple[str | None, str | None] = (None, None)
    for row_idx in range(3, ws.max_row + 1):
        title = ws.cell(row_idx, 3).value
        amount = ws.cell(row_idx, 4).value
        group = ws.cell(row_idx, 2).value
        if title is None and amount is None:
            continue
        if group is not None:
            last_date = normalize_date_label(group)
        entry_date, date_label = last_date
        amount_value, amount_expr = parse_amount(amount, cached_ws.cell(row_idx, 4).value)
        planned = date_label in {"나갈 돈", "카드 정기결제"}
        entries.append(
            {
                "book_section": "current",
                "entry_kind": "planned" if planned else "expense",
                "entry_date": entry_date,
                "date_label": date_label,
                "group_label": None if entry_date else date_label,
                "title": "" if title is None else str(title),
                "usage_place": None,
                "usage_item": None,
                "amount_value": amount_value,
                "amount_expr": amount_expr,
                "sort_order": row_idx,
                "due_day": _extract_due_day(str(title)) if planned and title is not None else None,
                "confirmed_at": None,
                "spending_category": None,
            }
        )
    return entries


def _extract_due_day(text: str) -> int | None:
    import re

    match = re.search(r"매월\s*(\d{1,2})\s*일", text)
    if not match:
        return None
    day = int(match.group(1))
    return day if 1 <= day <= 31 else None


def _parse_current_panels(ws: Any, cached_ws: Any) -> list[dict[str, Any]]:
    panel_ranges = {
        "fixed": range(4, 8),
        "frozen": range(11, 14),
        "claim": range(18, 33),
        "settlement": range(47, 56),
    }
    panels: list[dict[str, Any]] = []
    for panel_type, rows in panel_ranges.items():
        for row_idx in rows:
            title = ws.cell(row_idx, 6).value
            amount = ws.cell(row_idx, 7).value
            if title is None and amount is None:
                continue
            amount_value, amount_expr = parse_amount(amount, cached_ws.cell(row_idx, 7).value)
            panels.append(
                {
                    "month": _detect_current_month(ws),
                    "panel_type": panel_type,
                    "title": "" if title is None else str(title),
                    "amount_value": amount_value,
                    "amount_expr": amount_expr,
                    "sort_order": row_idx,
                }
            )
    return panels


def _detect_current_month(ws: Any) -> str:
    for row_idx in range(3, ws.max_row + 1):
        entry_date, _ = normalize_date_label(ws.cell(row_idx, 2).value)
        if entry_date:
            return entry_date[:7]
    return date.today().strftime("%Y-%m")


def _parse_current_settings(ws: Any, cached_ws: Any) -> dict[str, str]:
    base = "400000"
    formula = ws["J9"].value
    if isinstance(formula, str) and formula.startswith("="):
        expression = formula[1:]
        first_token = expression.split("-", 1)[0].strip()
        if _to_float(first_token) is not None:
            base = str(int(_to_float(first_token) or 0))
    return {
        "base_next_month_liquidity": base,
        "interest_expense": str(_to_float(cached_ws["J5"].value) or 0),
        "liquidity_status": str(_to_float(cached_ws["J7"].value) or 0),
    }


def _parse_labels(current: Any, archive: Any) -> dict[str, str]:
    return {
        "current_header_date": _label(current["B2"].value, "날짜"),
        "current_header_title": _label(current["C2"].value, "적요"),
        "current_header_amount": _label(current["D2"].value, "금액"),
        "archive_header_date": _label(archive["B2"].value, "날짜"),
        "archive_header_title": _label(archive["C2"].value, "적요"),
        "archive_header_amount": _label(archive["D2"].value, "금액"),
        "panel_fixed_title": _label(current["F2"].value, "현금성 고정지출"),
        "panel_frozen_title": _label(current["F9"].value, "동결"),
        "panel_claim_title": _label(current["F16"].value, "청구"),
        "panel_settlement_title": _label(current["F45"].value, "타인정산"),
        "panel_header_title": _label(current["F3"].value, "적요"),
        "panel_header_amount": _label(current["G3"].value, "금액"),
        "summary_title": _label(current["I2"].value, "요약"),
        "summary_card_total_label": _label(current["I3"].value, "카드대금"),
        "summary_transfer_or_deposit_label": _label(current["I4"].value, "송금/예치"),
        "summary_interest_expense_label": _label(current["I5"].value, "이자지출"),
        "summary_frozen_asset_label": _label(current["I6"].value, "동결자산"),
        "summary_liquidity_status_label": _label(current["I7"].value, "유동성 현황"),
        "summary_next_month_liquidity_label": _label(current["I9"].value, "익월 유동성"),
    }


def _label(value: Any, fallback: str) -> str:
    if value is None:
        return fallback
    return str(value)


def _parse_archive_rows(ws: Any, cached_ws: Any) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for row_idx in range(3, ws.max_row + 1):
        b_value = _hard_cell_value(ws, cached_ws, row_idx, 2)
        c_value = _hard_cell_value(ws, cached_ws, row_idx, 3)
        d_value = _to_float(_hard_cell_value(ws, cached_ws, row_idx, 4))
        e_value = _hard_cell_value(ws, cached_ws, row_idx, 5)
        f_value = _hard_cell_value(ws, cached_ws, row_idx, 6)
        if b_value is None and c_value is None and d_value is None and e_value is None and f_value is None:
            continue
        rows.append(
            {
                "source_row": row_idx,
                "b_value": _as_text(b_value),
                "c_value": _as_text(c_value),
                "d_value": d_value,
                "e_value": _as_text(e_value),
                "f_value": _as_text(f_value),
                "merge_down": _merge_down_count(ws, row_idx, 2),
                "sort_order": row_idx,
            }
        )
    return rows


def _hard_cell_value(ws: Any, cached_ws: Any, row_idx: int, col_idx: int) -> Any:
    cell = ws.cell(row_idx, col_idx)
    if isinstance(cell, MergedCell):
        return None
    value = cell.value
    if isinstance(value, str) and value.startswith("="):
        return cached_ws.cell(row_idx, col_idx).value
    return value


def _merge_down_count(ws: Any, row_idx: int, col_idx: int) -> int:
    coordinate = ws.cell(row_idx, col_idx).coordinate
    for merged_range in ws.merged_cells.ranges:
        if (
            coordinate in merged_range
            and merged_range.min_row == row_idx
            and merged_range.min_col == col_idx
        ):
            return merged_range.max_row - merged_range.min_row
    return 0


def _as_text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return f"{value.date():%Y.%m.%d}."
    if isinstance(value, date):
        return f"{value:%Y.%m.%d}."
    return str(value)


def export_workbook(
    hard_archive_rows: list[dict[str, Any]],
    archive_entries: list[dict[str, Any]],
    current_entries: list[dict[str, Any]],
    panels: list[dict[str, Any]],
    labels: dict[str, str],
    settings: dict[str, str] | None,
    output_path: Path,
    template_path: Path | None = None,
) -> Path:
    if template_path and template_path.exists():
        wb = load_workbook(template_path)
        current = wb[CURRENT_SHEET]
        archive = wb[ARCHIVE_SHEET]
        _clear_current_sheet_values(current)
        preserve_hard_archive = True
    else:
        wb = Workbook()
        current = wb.active
        current.title = CURRENT_SHEET
        archive = wb.create_sheet(ARCHIVE_SHEET)
        preserve_hard_archive = False

    _write_current_sheet(
        current,
        current_entries,
        panels,
        labels,
        settings or {},
        preserve_dimensions=preserve_hard_archive,
    )
    _write_archive_sheet(
        archive,
        hard_archive_rows,
        archive_entries,
        labels,
        preserve_hard_rows=preserve_hard_archive,
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def _clear_current_sheet_values(ws: Any) -> None:
    _reset_dynamic_panel_merges(ws)
    _reset_main_entry_merges(ws)
    ranges = [
        (3, 300, 2, 4),
        (2, 300, 6, 7),
        (2, 9, 9, 10),
    ]
    for min_row, max_row, min_col, max_col in ranges:
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row, col)
                if not isinstance(cell, MergedCell):
                    cell.value = None


def _reset_dynamic_panel_merges(ws: Any) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_col == 6 and merged_range.max_col == 7 and merged_range.min_row >= 9:
            ws.unmerge_cells(str(merged_range))


def _reset_main_entry_merges(ws: Any) -> None:
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_col == 2 and merged_range.max_col == 2 and merged_range.min_row >= 3:
            ws.unmerge_cells(str(merged_range))


def _write_current_sheet(
    ws: Any,
    entries: list[dict[str, Any]],
    panels: list[dict[str, Any]],
    labels: dict[str, str],
    settings: dict[str, str],
    preserve_dimensions: bool = False,
) -> None:
    _write_header(ws, "B2", _get_label(labels, "current_header_date", "날짜"))
    _write_header(ws, "C2", _get_label(labels, "current_header_title", "적요"))
    _write_header(ws, "D2", _get_label(labels, "current_header_amount", "금액"))
    _write_header(ws, "F2", _get_label(labels, "panel_fixed_title", "현금성 고정지출"))
    _write_header(ws, "I2", _get_label(labels, "summary_title", "요약"))

    current_first, current_last = _write_current_entries_table(ws, entries)

    _write_fixed_panel(ws, panels, labels)
    panel_layout = _write_dynamic_panels(ws, panels, labels)

    ws["I3"] = _get_label(labels, "summary_card_total_label", "카드대금")
    ws["J3"] = f"=SUM(D{current_first}:D{current_last})" if current_last >= current_first else 0
    ws["I4"] = _get_label(labels, "summary_transfer_or_deposit_label", "송금/예치")
    fixed_first, fixed_last = panel_layout["fixed"]
    ws["J4"] = f"=SUM(G{fixed_first}:G{fixed_last})" if fixed_last >= fixed_first else 0
    ws["I5"] = _get_label(labels, "summary_interest_expense_label", "이자지출")
    ws["J5"] = _setting_number(settings, "interest_expense", 0)
    ws["I6"] = _get_label(labels, "summary_frozen_asset_label", "동결자산")
    frozen_first, frozen_last = panel_layout["frozen"]
    ws["J6"] = f"=SUM(G{frozen_first}:G{frozen_last})" if frozen_last >= frozen_first else 0
    ws["I7"] = _get_label(labels, "summary_liquidity_status_label", "유동성 현황")
    ws["J7"] = _setting_number(settings, "liquidity_status", 0)
    ws["I9"] = _get_label(labels, "summary_next_month_liquidity_label", "익월 유동성")
    ws["J9"] = f"={_setting_number(settings, 'base_next_month_liquidity', 400000):g}-SUM(J3:J6)+J7"

    if not preserve_dimensions:
        ws.column_dimensions["B"].width = 13
        ws.column_dimensions["C"].width = 90
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["F"].width = 42
        ws.column_dimensions["G"].width = 14
        ws.column_dimensions["I"].width = 18
        ws.column_dimensions["J"].width = 18


def _write_current_entries_table(ws: Any, entries: list[dict[str, Any]]) -> tuple[int, int]:
    start_row = 3
    occupied_rows: set[int] = set()
    group_ranges: list[tuple[int, int]] = []
    group_start: int | None = None
    previous_label: str | None = None

    for offset, entry in enumerate(entries):
        row = start_row + offset
        label = entry.get("date_label") or entry.get("group_label")
        is_planned = entry.get("entry_kind") == "planned"
        starts_group = offset == 0 or label != previous_label
        source_row = 3 if is_planned and starts_group else 4 if is_planned else 9 if starts_group else 10

        _copy_row_style(ws, source_row, row, 2, 4)
        _write_cell_value(ws, row, 2, label if starts_group else None)
        _write_cell_value(ws, row, 3, entry.get("title"))
        _write_amount(ws.cell(row, 4), entry.get("amount_value"), entry.get("amount_expr"))
        occupied_rows.add(row)

        if starts_group:
            if group_start is not None and row - 1 > group_start:
                ws.merge_cells(start_row=group_start, start_column=2, end_row=row - 1, end_column=2)
                group_ranges.append((group_start, row - 1))
            elif group_start is not None:
                group_ranges.append((group_start, group_start))
            group_start = row
        previous_label = label

    if group_start is not None:
        last_row = start_row + len(entries) - 1
        if last_row > group_start:
            ws.merge_cells(start_row=group_start, start_column=2, end_row=last_row, end_column=2)
            group_ranges.append((group_start, last_row))
        else:
            group_ranges.append((group_start, group_start))

    for first_row, last_row in group_ranges:
        _style_current_date_group(ws, first_row, last_row)

    _clear_unused_main_rows(ws, occupied_rows, min_row=start_row, max_row=300)
    return (start_row, start_row + len(entries) - 1) if entries else (start_row, start_row - 1)


def _write_fixed_panel(ws: Any, panels: list[dict[str, Any]], labels: dict[str, str]) -> None:
    _write_panel_table(
        ws=ws,
        panel_type="fixed",
        title_row=2,
        rows=[panel for panel in panels if panel.get("panel_type") == "fixed"],
        labels=labels,
        title_label=_get_label(labels, "panel_fixed_title", "현금성 고정지출"),
        style_source_rows=(2, 3, 4),
    )


def _write_dynamic_panels(ws: Any, panels: list[dict[str, Any]], labels: dict[str, str]) -> dict[str, tuple[int, int]]:
    data_ranges: dict[str, tuple[int, int]] = {}
    fixed_rows = [panel for panel in panels if panel.get("panel_type") == "fixed"]
    fixed_end = 4 + max(len(fixed_rows), 1) - 1
    data_ranges["fixed"] = (4, 4 + len(fixed_rows) - 1) if fixed_rows else (4, 3)
    layout = [
        ("frozen", "panel_frozen_title", "동결", (2, 3, 4)),
        ("claim", "panel_claim_title", "청구", (2, 3, 4)),
        ("settlement", "panel_settlement_title", "타인정산", (2, 3, 4)),
    ]
    title_row = max(10, fixed_end + 3)
    clear_start_row = fixed_end + 1
    occupied_rows: set[int] = set()
    for panel_type, label_key, fallback, style_source_rows in layout:
        rows = [panel for panel in panels if panel.get("panel_type") == panel_type]
        end_row, used_rows, data_range = _write_panel_table(
            ws=ws,
            panel_type=panel_type,
            title_row=title_row,
            rows=rows,
            labels=labels,
            title_label=_get_label(labels, label_key, fallback),
            style_source_rows=style_source_rows,
        )
        occupied_rows.update(used_rows)
        data_ranges[panel_type] = data_range
        title_row = end_row + 3
    _clear_unused_panel_rows(ws, occupied_rows, min_row=clear_start_row, max_row=max(55, title_row))
    return data_ranges


def _write_panel_table(
    ws: Any,
    panel_type: str,
    title_row: int,
    rows: list[dict[str, Any]],
    labels: dict[str, str],
    title_label: str,
    style_source_rows: tuple[int, int, int],
) -> tuple[int, set[int], tuple[int, int]]:
    header_row = title_row + 1
    first_data_row = title_row + 2
    source_title_row, source_header_row, source_data_row = style_source_rows

    _copy_row_style(ws, source_title_row, title_row, 6, 7)
    _copy_row_style(ws, source_header_row, header_row, 6, 7)
    _write_cell_value(ws, title_row, 6, title_label)
    _write_cell_value(ws, title_row, 7, None)
    _write_cell_value(ws, header_row, 6, _get_label(labels, "panel_header_title", "적요"))
    _write_cell_value(ws, header_row, 7, _get_label(labels, "panel_header_amount", "금액"))

    if panel_type != "fixed":
        _ensure_merged(ws, title_row, 6, title_row, 7)

    used_rows = {title_row, header_row}
    for offset, panel in enumerate(rows):
        row = first_data_row + offset
        used_rows.add(row)
        _copy_row_style(ws, source_data_row, row, 6, 7)
        _write_cell_value(ws, row, 6, panel.get("title"))
        _write_amount(ws.cell(row, 7), panel.get("amount_value"), panel.get("amount_expr"))

    if rows:
        data_range = (first_data_row, first_data_row + len(rows) - 1)
        end_row = data_range[1]
    else:
        data_range = (first_data_row, first_data_row - 1)
        end_row = header_row
    return end_row, used_rows, data_range


def _write_archive_sheet(
    ws: Any,
    hard_rows: list[dict[str, Any]],
    entries: list[dict[str, Any]],
    labels: dict[str, str],
    preserve_hard_rows: bool = False,
) -> None:
    _write_header(ws, "B2", _get_label(labels, "archive_header_date", "날짜"))
    _write_header(ws, "C2", _get_label(labels, "archive_header_title", "적요"))
    _write_header(ws, "D2", _get_label(labels, "archive_header_amount", "금액"))

    start_row = 3
    if not preserve_hard_rows:
        for offset, hard_row in enumerate(hard_rows):
            row = hard_row.get("source_row") or (start_row + offset)
            if not isinstance(ws.cell(row, 2), MergedCell):
                ws.cell(row, 2).value = hard_row.get("b_value")
            ws.cell(row, 3).value = hard_row.get("c_value")
            ws.cell(row, 4).value = hard_row.get("d_value")
            ws.cell(row, 5).value = hard_row.get("e_value")
            ws.cell(row, 6).value = hard_row.get("f_value")
            merge_down = hard_row.get("merge_down") or 0
            if merge_down:
                ws.merge_cells(start_row=row, start_column=2, end_row=row + merge_down, end_column=2)

    start_row = _next_row_after_hard_archive(start_row, hard_rows)
    date_style_row, continuation_style_row = _archive_append_style_rows(hard_rows)
    for offset, entry in enumerate(entries):
        row = start_row + offset
        starts_group = _starts_new_label(entries, offset)
        _copy_row_style(ws, date_style_row if starts_group else continuation_style_row, row, 2, 6)
        if starts_group:
            ws.cell(row, 2).value = entry.get("date_label") or entry.get("group_label")
        ws.cell(row, 3).value = entry.get("title")
        _write_amount(ws.cell(row, 4), entry.get("amount_value"), entry.get("amount_expr"))
        _write_amount(ws.cell(row, 5), entry.get("aux_amount_value"), entry.get("aux_amount_expr"))
        ws.cell(row, 6).value = entry.get("extra_value")

    _merge_archive_dates(ws, entries, start_row)
    _apply_previous_month_archive_fill(ws, entries, start_row)
    if not preserve_hard_rows:
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 92
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 13
        ws.column_dimensions["F"].width = 16


def _next_row_after_hard_archive(start_row: int, hard_rows: list[dict[str, Any]]) -> int:
    next_row = start_row
    for offset, hard_row in enumerate(hard_rows):
        row = hard_row.get("source_row") or (start_row + offset)
        merge_down = hard_row.get("merge_down") or 0
        next_row = max(next_row, row + merge_down + 1)
    return next_row


def _archive_append_style_rows(hard_rows: list[dict[str, Any]]) -> tuple[int, int]:
    if not hard_rows:
        return 3, 4
    sorted_rows = sorted(hard_rows, key=lambda row: row.get("source_row") or row.get("sort_order") or 0)
    date_style_row = None
    continuation_style_row = None
    for row in reversed(sorted_rows):
        source_row = row.get("source_row")
        if not source_row:
            continue
        if date_style_row is None and row.get("b_value") is not None:
            date_style_row = int(source_row)
        if continuation_style_row is None and row.get("b_value") is None and row.get("c_value") is not None:
            continuation_style_row = int(source_row)
        if date_style_row is not None and continuation_style_row is not None:
            break
    if date_style_row is None:
        date_style_row = int(sorted_rows[-1].get("source_row") or 3)
    if continuation_style_row is None:
        continuation_style_row = date_style_row
    return date_style_row, continuation_style_row


def _starts_new_label(entries: list[dict[str, Any]], offset: int) -> bool:
    if offset == 0:
        return True
    current = entries[offset].get("date_label") or entries[offset].get("group_label")
    previous = entries[offset - 1].get("date_label") or entries[offset - 1].get("group_label")
    return current != previous


def _merge_archive_dates(ws: Any, entries: list[dict[str, Any]], start_row: int) -> None:
    group_start = start_row
    previous = None
    for offset, entry in enumerate(entries + [{"date_label": "__sentinel__"}]):
        label = entry.get("date_label") or entry.get("group_label")
        row = start_row + offset
        if previous is None:
            previous = label
            continue
        if label != previous:
            if row - 1 > group_start and previous:
                ws.merge_cells(start_row=group_start, start_column=2, end_row=row - 1, end_column=2)
            group_start = row
            previous = label


def _style_current_date_group(ws: Any, first_row: int, last_row: int) -> None:
    cell = ws.cell(first_row, 2)
    if isinstance(cell, MergedCell):
        return
    source_row = 3 if first_row <= 8 else 9
    source = ws.cell(source_row, 2)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = _complete_border(source.border)
    for row in range(first_row + 1, last_row + 1):
        target = ws.cell(row, 2)
        if isinstance(target, MergedCell):
            continue
        target.alignment = Alignment(horizontal="center", vertical="center")
        target.border = _complete_border(source.border)


def _complete_border(border: Border) -> Border:
    fallback = Side(style="thin", color="000000")
    return Border(
        left=_border_side_or_fallback(border.left, fallback),
        right=_border_side_or_fallback(border.right, fallback),
        top=_border_side_or_fallback(border.top, fallback),
        bottom=_border_side_or_fallback(border.bottom, fallback),
    )


def _border_side_or_fallback(side: Side, fallback: Side) -> Side:
    return copy(side) if side and side.style else copy(fallback)


def _apply_previous_month_archive_fill(ws: Any, entries: list[dict[str, Any]], start_row: int) -> None:
    months = [
        str(entry.get("entry_date"))[:7]
        for entry in entries
        if entry.get("entry_date")
    ]
    if not months:
        return
    latest_month = max(months)
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFFF00")
    for offset, entry in enumerate(entries):
        if str(entry.get("entry_date") or "")[:7] != latest_month:
            continue
        row = start_row + offset
        for col in range(3, 5):
            ws.cell(row, col).fill = copy(yellow_fill)


def _write_amount(cell: Any, amount_value: Any, amount_expr: Any) -> None:
    if amount_expr:
        cell.value = amount_expr
    else:
        cell.value = amount_value


def _write_header(ws: Any, coordinate: str, value: str) -> None:
    cell = ws[coordinate]
    cell.value = value


def _write_cell_value(ws: Any, row: int, col: int, value: Any) -> None:
    cell = ws.cell(row, col)
    if not isinstance(cell, MergedCell):
        cell.value = value


def _copy_row_style(ws: Any, source_row: int, target_row: int, min_col: int, max_col: int) -> None:
    if source_row == target_row:
        return
    for col in range(min_col, max_col + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if isinstance(target, MergedCell):
            continue
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.protection = copy(source.protection)


def _clear_unused_panel_rows(ws: Any, occupied_rows: set[int], min_row: int, max_row: int) -> None:
    for row in range(min_row, max_row + 1):
        if row in occupied_rows:
            continue
        for col in range(6, 8):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell._style = copy(ws.cell(1, 1)._style)


def _clear_unused_main_rows(ws: Any, occupied_rows: set[int], min_row: int, max_row: int) -> None:
    for row in range(min_row, max_row + 1):
        if row in occupied_rows:
            continue
        for col in range(2, 5):
            cell = ws.cell(row, col)
            if isinstance(cell, MergedCell):
                continue
            cell.value = None
            cell._style = copy(ws.cell(1, 1)._style)


def _ensure_merged(ws: Any, min_row: int, min_col: int, max_row: int, max_col: int) -> None:
    target = f"{ws.cell(min_row, min_col).coordinate}:{ws.cell(max_row, max_col).coordinate}"
    if target not in {str(merged_range) for merged_range in ws.merged_cells.ranges}:
        ws.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)


def _get_label(labels: dict[str, str], key: str, fallback: str) -> str:
    return labels.get(key) or fallback


def _setting_number(settings: dict[str, str], key: str, fallback: float) -> float:
    try:
        return float(settings.get(key, fallback))
    except (TypeError, ValueError):
        return fallback
